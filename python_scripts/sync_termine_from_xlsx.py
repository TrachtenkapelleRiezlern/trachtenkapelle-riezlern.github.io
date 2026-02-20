#!/usr/bin/env python3
"""
sync_termine_from_xlsx.py
──────────────────────────
Liest Termine aus einer lokalen Excel-Datei (.xlsx) im Repo
und generiert/aktualisiert Termin-Ordner mit meta.json.

- Nur Zeilen mit "Ja" in "Öffentlich" werden verarbeitet.
- Termin-Ordner die NICHT (mehr) in der Excel stehen werden gelöscht.

Voraussetzungen:  pip install openpyxl
Verwendung:
    python3 sync_termine_from_xlsx.py
    python3 sync_termine_from_xlsx.py --file data/termine.xlsx
    python3 sync_termine_from_xlsx.py --sheet "Tabelle1"
    python3 sync_termine_from_xlsx.py --dry-run
"""

import sys, json, re, unicodedata, argparse, shutil
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Fehlendes Paket: pip install openpyxl")
    sys.exit(1)

# ── Konfiguration ─────────────────────────────────────────────────────────────
DEFAULT_XLSX  = "data/termine.xlsx"
DEFAULT_SHEET = None
TERMINE_DIR   = Path(__file__).parent.parent / "Termine"
INDEX_FILE    = TERMINE_DIR / "index.json"
FOLDER_RE     = re.compile(r"^\d{4}_\d{2}_\d{2}_")  # gültige Termin-Ordner

# ── Spalten ───────────────────────────────────────────────────────────────────
COL_OEFFENTLICH   = "öffentlich"
COL_START_TIME    = "start time"
COL_TITLE         = "title"
COL_VERANSTALTUNG = "veranstaltung"
COL_LOCATION      = "location"
COL_BESCHREIBUNG  = "beschreibung"
COL_DESCRIPTION   = "description"
COL_EINTRITT      = "eintritt"
COL_KATEGORIE     = "kategorie"
COL_ORCHESTER     = "orchester"
COL_END_TIME      = "end time"
COL_ID            = "id"

KATEGORIE_MAP = {
    "konzert": "konzert", "konzerte": "konzert",
    "prozession": "prozession", "prozessionen": "prozession",
    "messe": "prozession", "gottesdienst": "prozession", "kirchlich": "prozession",
    "fest": "fest", "feste": "fest", "festakt": "fest",
    "auswärts": "auswaerts", "auswaerts": "auswaerts", "auswärtsspiel": "auswaerts",
    "sonstiges": "sonstiges", "sonstige": "sonstiges",
}

# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

def slugify(text):
    text = unicodedata.normalize("NFKD", str(text))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text).strip()
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:40].rstrip("_")


def parse_cell_datetime(value):
    """
    Parst Datum+Uhrzeit aus Excel-Zelle.
    Unterstützt u.a. "Sa. 25.04.26 20:00", "25.04.2026", "2026-04-25 20:00"
    """
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        uhrzeit = value.strftime("%H:%M")
        return value.strftime("%Y-%m-%d"), (None if uhrzeit == "00:00" else uhrzeit)
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d"), None

    raw = str(value).strip()
    if not raw:
        return None, None

    # "Sa. " / "Do. " etc. entfernen
    cleaned = re.sub(r'^[A-Za-z\u00c4\u00d6\u00dc\u00e4\u00f6\u00fc]+\.\s*', '', raw)
    # Zweistelliges Jahr: dd.mm.yy → dd.mm.20yy
    cleaned = re.sub(
        r'^(\d{2}\.\d{2}\.)(\d{2})(\s|$)',
        lambda m: f"{m.group(1)}20{m.group(2)}{m.group(3)}",
        cleaned
    ).strip()

    formats_mit_zeit = [
        "%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
        "%m/%d/%Y %H:%M", "%m/%d/%Y %I:%M %p",
    ]
    formats_nur_datum = ["%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"]

    for candidate in ([cleaned, raw] if cleaned != raw else [raw]):
        for fmt in formats_mit_zeit:
            try:
                dt = datetime.strptime(candidate, fmt)
                uhrzeit = dt.strftime("%H:%M")
                return dt.strftime("%Y-%m-%d"), (None if uhrzeit == "00:00" else uhrzeit)
            except ValueError:
                pass
        for fmt in formats_nur_datum:
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d"), None
            except ValueError:
                pass
    return None, None


def cell_str(value):
    return "" if value is None else str(value).strip()


def normalize_kategorie(raw):
    return KATEGORIE_MAP.get(raw.strip().lower(), "sonstiges")


def is_oeffentlich(value):
    return cell_str(value).lower() in ("ja", "yes", "true", "1", "x", "✓", "✔")


def load_xlsx(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Blatt '{sheet_name}' nicht gefunden. Verfügbar: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active
        print(f"Tabellenblatt: '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [cell_str(h).lower() for h in rows[0]]
    missing = [c for c in [COL_OEFFENTLICH, COL_START_TIME] if c not in headers]
    if missing:
        print(f"Fehler – Pflicht-Spalten fehlen: {missing}")
        print(f"Gefunden: {[h for h in headers if h]}")
        sys.exit(1)

    dicts = []
    for row in rows[1:]:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        if any(v is not None and str(v).strip() for v in d.values()):
            dicts.append(d)
    return headers, dicts


# ── Hauptlogik ────────────────────────────────────────────────────────────────

def sync(xlsx_path, sheet_name=None, dry_run=False):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"Datei nicht gefunden: {xlsx_path}")
        sys.exit(1)

    print(f"Excel-Datei : {xlsx_path}")
    if dry_run:
        print("DRY RUN – keine Dateien werden geschrieben/gelöscht")
    print()

    headers, dicts = load_xlsx(xlsx_path, sheet_name)
    print(f"Spalten     : {[h for h in headers if h]}")
    print(f"Datenzeilen : {len(dicts)}")
    print()

    TERMINE_DIR.mkdir(parents=True, exist_ok=True)

    created = updated = skipped = errors = deleted = 0
    index_entries = []
    synced_ordner = set()  # alle Ordner die aus dem Sheet kommen

    # ── Zeilen verarbeiten ────────────────────────────────────────────────────
    for i, row in enumerate(dicts, start=2):

        if not is_oeffentlich(row.get(COL_OEFFENTLICH)):
            skipped += 1
            continue

        datum_iso, uhrzeit = parse_cell_datetime(row.get(COL_START_TIME))
        if not datum_iso:
            print(f"  Zeile {i:4d}: Ungültiges Datum '{cell_str(row.get(COL_START_TIME))}' – übersprungen")
            errors += 1
            continue

        titel = cell_str(row.get(COL_TITLE)) or cell_str(row.get(COL_VERANSTALTUNG))
        if not titel:
            print(f"  Zeile {i:4d}: Kein Titel – übersprungen")
            errors += 1
            continue

        ort          = cell_str(row.get(COL_LOCATION)) or None
        beschreibung = cell_str(row.get(COL_BESCHREIBUNG)) or cell_str(row.get(COL_DESCRIPTION)) or None
        kat_raw      = cell_str(row.get(COL_KATEGORIE))
        kategorie    = normalize_kategorie(kat_raw) if kat_raw else "sonstiges"
        eintritt     = cell_str(row.get(COL_EINTRITT)) or None
        orchester    = cell_str(row.get(COL_ORCHESTER)) or None
        cal_id       = cell_str(row.get(COL_ID)) or None
        _, end_uhrzeit = parse_cell_datetime(row.get(COL_END_TIME))

        ordner_name = f"{datum_iso.replace('-', '_')}_{slugify(titel)}"
        ordner_path = TERMINE_DIR / ordner_name
        meta_path   = ordner_path / "meta.json"
        synced_ordner.add(ordner_name)

        meta = {"titel": titel, "datum": datum_iso}
        if uhrzeit:      meta["uhrzeit"]       = uhrzeit
        if end_uhrzeit:  meta["uhrzeit_ende"]  = end_uhrzeit
        if ort:          meta["ort"]           = ort
        meta["kategorie"] = kategorie
        if beschreibung: meta["beschreibung"]  = beschreibung
        if eintritt:     meta["eintritt"]      = eintritt
        if orchester:    meta["orchester"]     = orchester
        if cal_id:       meta["google_cal_id"] = cal_id
        meta["bilder"]   = []
        meta["featured"] = False

        # Bilder + featured aus bestehendem meta.json erhalten
        if ordner_path.exists() and meta_path.exists():
            try:
                old = json.loads(meta_path.read_text(encoding="utf-8"))
                if old.get("bilder"):   meta["bilder"]   = old["bilder"]
                if old.get("featured"): meta["featured"] = old["featured"]
            except Exception:
                pass

        is_new = not ordner_path.exists()
        uhr_str = f" {uhrzeit}" if uhrzeit else ""
        ort_str = f" @ {ort}" if ort else ""
        print(f"  {'+'if is_new else '~'}  {datum_iso}{uhr_str}{ort_str}  {titel}  [{kategorie}]")

        if not dry_run:
            ordner_path.mkdir(parents=True, exist_ok=True)
            meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        if is_new: created += 1
        else:      updated += 1

        index_entries.append({"ordner": ordner_name, "datum": datum_iso, "titel": titel})

    # ── Veraltete Ordner löschen ──────────────────────────────────────────────
    print()
    for folder in sorted(TERMINE_DIR.iterdir()):
        if not folder.is_dir() or not FOLDER_RE.match(folder.name):
            continue
        if folder.name not in synced_ordner:
            print(f"  ✗  {folder.name}  ← nicht mehr im Sheet, wird gelöscht")
            if not dry_run:
                shutil.rmtree(folder)
            deleted += 1

    # ── index.json ────────────────────────────────────────────────────────────
    index_entries.sort(key=lambda e: e["datum"])
    if not dry_run:
        INDEX_FILE.write_text(
            json.dumps(index_entries, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    print()
    print("─" * 52)
    print(f"  {created:3d}  neu erstellt")
    print(f"  {updated:3d}  aktualisiert")
    print(f"  {deleted:3d}  gelöscht      (nicht mehr im Sheet)")
    print(f"  {skipped:3d}  übersprungen  (Öffentlich ≠ Ja)")
    print(f"  {errors:3d}  Fehler        (fehlendes Datum / Titel)")
    print(f"  {len(index_entries):3d}  Einträge in Termine/index.json")
    if dry_run:
        print("\n  DRY RUN – keine Dateien wurden geändert.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", "-f", default=DEFAULT_XLSX)
    parser.add_argument("--sheet", "-s", default=DEFAULT_SHEET)
    parser.add_argument("--dry-run", "-n", action="store_true")
    args = parser.parse_args()
    sync(args.file, args.sheet, args.dry_run)